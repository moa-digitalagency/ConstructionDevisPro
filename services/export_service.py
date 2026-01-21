from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import io
from datetime import datetime


class ExportService:
    def __init__(self, company):
        self.company = company
        self.branding = company.branding
        self.tax_profile = company.tax_profile
    
    def generate_pdf(self, quote, version):
        buffer = io.BytesIO()
        
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor(self.branding.primary_color or '#1a56db')
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=12,
            spaceAfter=10,
            textColor=colors.HexColor(self.branding.primary_color or '#1a56db')
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=5
        )
        
        elements = []
        
        elements.append(Paragraph(f"DEVIS N° {quote.reference}", title_style))
        elements.append(Paragraph(f"Version {version.version_number}", styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
        
        company_info = f"""
        <b>{self.branding.legal_name or self.company.name}</b><br/>
        {self.branding.address or ''}<br/>
        Tél: {self.branding.phone or ''}<br/>
        Email: {self.branding.email or ''}<br/>
        {f"RC: {self.branding.registration_number}" if self.branding.registration_number else ''}<br/>
        {f"IF: {self.branding.tax_id}" if self.branding.tax_id else ''}
        """
        
        project = quote.project
        client_info = f"""
        <b>Client:</b><br/>
        {project.client_name or 'Non renseigné'}<br/>
        {project.client_address or ''}<br/>
        {project.client_phone or ''}<br/>
        {project.client_email or ''}
        """
        
        info_data = [[
            Paragraph(company_info, normal_style),
            Paragraph(client_info, normal_style)
        ]]
        
        info_table = Table(info_data, colWidths=[9*cm, 9*cm])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*cm))
        
        elements.append(Paragraph(f"<b>Projet:</b> {project.name}", normal_style))
        elements.append(Paragraph(f"<b>Type:</b> {project.project_type.value.capitalize()}", normal_style))
        elements.append(Paragraph(f"<b>Date:</b> {datetime.now().strftime('%d/%m/%Y')}", normal_style))
        if quote.valid_until:
            elements.append(Paragraph(f"<b>Valide jusqu'au:</b> {quote.valid_until.strftime('%d/%m/%Y')}", normal_style))
        elements.append(Spacer(1, 0.5*cm))
        
        elements.append(Paragraph("DÉTAIL DES PRESTATIONS", header_style))
        
        table_data = [['Désignation', 'Unité', 'Qté', 'P.U. HT', 'Total HT']]
        
        current_category = None
        for line in version.lines.order_by('sort_order'):
            if line.category != current_category:
                table_data.append([Paragraph(f"<b>{line.category}</b>", normal_style), '', '', '', ''])
                current_category = line.category
            
            table_data.append([
                Paragraph(line.designation, normal_style),
                line.unit,
                f"{float(line.quantity):.2f}",
                f"{float(line.unit_price):,.2f}",
                f"{float(line.total_price):,.2f}"
            ])
        
        lines_table = Table(table_data, colWidths=[8*cm, 2*cm, 2*cm, 3*cm, 3*cm])
        lines_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.branding.primary_color or '#1a56db')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(lines_table)
        elements.append(Spacer(1, 0.5*cm))
        
        currency = self.company.currency
        totals_data = [
            ['', '', '', 'Total HT:', f"{float(version.subtotal_ht):,.2f} {currency}"],
            ['', '', '', f'TVA ({float(version.vat_rate)}%):', f"{float(version.vat_amount):,.2f} {currency}"],
            ['', '', '', 'TOTAL TTC:', f"{float(version.total_ttc):,.2f} {currency}"],
        ]
        
        totals_table = Table(totals_data, colWidths=[6*cm, 2*cm, 2*cm, 4*cm, 4*cm])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (3, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('LINEABOVE', (3, -1), (-1, -1), 1, colors.black),
        ]))
        elements.append(totals_table)
        elements.append(Spacer(1, 1*cm))
        
        if version.assumptions.count() > 0:
            elements.append(Paragraph("HYPOTHÈSES ET CONDITIONS", header_style))
            for assumption in version.assumptions:
                elements.append(Paragraph(f"• {assumption.description}: {assumption.value}", normal_style))
            elements.append(Spacer(1, 0.5*cm))
        
        if self.branding.quote_terms:
            elements.append(Paragraph("CONDITIONS GÉNÉRALES", header_style))
            elements.append(Paragraph(self.branding.quote_terms, normal_style))
            elements.append(Spacer(1, 0.5*cm))
        
        if self.branding.bank_details:
            elements.append(Paragraph("COORDONNÉES BANCAIRES", header_style))
            elements.append(Paragraph(self.branding.bank_details, normal_style))
        
        if self.branding.quote_footer:
            elements.append(Spacer(1, 1*cm))
            elements.append(Paragraph(self.branding.quote_footer, normal_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        output_dir = os.path.join('static', 'exports', str(self.company.id))
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{quote.reference}_V{version.version_number}.pdf")
        
        with open(output_path, 'wb') as f:
            f.write(buffer.read())
        
        version.pdf_path = output_path
        from models import db
        db.session.commit()
        
        return output_path
    
    def generate_excel(self, quote, version):
        wb = Workbook()
        ws = wb.active
        ws.title = "DQE"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        ws['A1'] = f"DEVIS N° {quote.reference} - Version {version.version_number}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = f"Projet: {quote.project.name}"
        ws['A3'] = f"Client: {quote.project.client_name or 'Non renseigné'}"
        ws['A4'] = f"Date: {datetime.now().strftime('%d/%m/%Y')}"
        
        headers = ['Catégorie', 'Désignation', 'Unité', 'Quantité', 'Prix Unitaire HT', 'Total HT']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        row = 7
        for line in version.lines.order_by('sort_order'):
            ws.cell(row=row, column=1, value=line.category).border = border
            ws.cell(row=row, column=2, value=line.designation).border = border
            ws.cell(row=row, column=3, value=line.unit).border = border
            ws.cell(row=row, column=4, value=float(line.quantity)).border = border
            ws.cell(row=row, column=5, value=float(line.unit_price)).border = border
            ws.cell(row=row, column=6, value=float(line.total_price)).border = border
            
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=5).number_format = '#,##0.00'
            ws.cell(row=row, column=6).number_format = '#,##0.00'
            
            row += 1
        
        row += 1
        ws.cell(row=row, column=5, value='Total HT:').font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(version.subtotal_ht)).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        
        row += 1
        ws.cell(row=row, column=5, value=f'TVA ({float(version.vat_rate)}%):')
        ws.cell(row=row, column=6, value=float(version.vat_amount))
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        
        row += 1
        ws.cell(row=row, column=5, value='Total TTC:').font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(version.total_ttc)).font = Font(bold=True)
        ws.cell(row=row, column=6).number_format = '#,##0.00'
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18
        
        if version.assumptions.count() > 0:
            ws_hyp = wb.create_sheet(title="Hypothèses")
            ws_hyp['A1'] = "HYPOTHÈSES DU DEVIS"
            ws_hyp['A1'].font = Font(bold=True, size=12)
            
            ws_hyp.cell(row=3, column=1, value="Catégorie").font = header_font
            ws_hyp.cell(row=3, column=1).fill = header_fill
            ws_hyp.cell(row=3, column=2, value="Description").font = header_font
            ws_hyp.cell(row=3, column=2).fill = header_fill
            ws_hyp.cell(row=3, column=3, value="Valeur").font = header_font
            ws_hyp.cell(row=3, column=3).fill = header_fill
            
            row = 4
            for assumption in version.assumptions:
                ws_hyp.cell(row=row, column=1, value=assumption.category)
                ws_hyp.cell(row=row, column=2, value=assumption.description)
                ws_hyp.cell(row=row, column=3, value=assumption.value)
                row += 1
            
            ws_hyp.column_dimensions['A'].width = 20
            ws_hyp.column_dimensions['B'].width = 50
            ws_hyp.column_dimensions['C'].width = 30
        
        output_dir = os.path.join('static', 'exports', str(self.company.id))
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, f"{quote.reference}_DQE_V{version.version_number}.xlsx")
        
        wb.save(output_path)
        
        return output_path
