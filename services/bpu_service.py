from models import db, BPULibrary, BPUArticle, CompanyBPUOverride, CompanyBPUArticle
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import io


class BPUService:
    def __init__(self, company):
        self.company = company
        self.library = BPULibrary.query.filter_by(
            country=company.country,
            is_active=True
        ).order_by(BPULibrary.version.desc()).first()
    
    def get_article_price(self, article, tier_code='STD'):
        override = CompanyBPUOverride.query.filter_by(
            company_id=self.company.id,
            article_id=article.id
        ).first()
        
        if override and override.is_disabled:
            return None
        
        if tier_code == 'ECO':
            if override and override.unit_price_eco:
                return override.unit_price_eco
            return article.unit_price_eco
        elif tier_code == 'PREM':
            if override and override.unit_price_premium:
                return override.unit_price_premium
            return article.unit_price_premium
        else:
            if override and override.unit_price_standard:
                return override.unit_price_standard
            return article.unit_price_standard
    
    def get_article_designation(self, article):
        override = CompanyBPUOverride.query.filter_by(
            company_id=self.company.id,
            article_id=article.id
        ).first()
        
        if override and override.designation_override:
            return override.designation_override
        return article.designation
    
    def export_to_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "BPU"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a56db", end_color="1a56db", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = ['Code', 'Catégorie', 'Sous-catégorie', 'Désignation', 'Unité', 
                   'Prix Éco', 'Prix Standard', 'Prix Premium', 'Personnalisé']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        if self.library:
            articles = BPUArticle.query.filter_by(library_id=self.library.id)\
                .order_by(BPUArticle.category, BPUArticle.sort_order).all()
            
            overrides = {o.article_id: o for o in self.company.bpu_overrides.all()}
            
            row = 2
            for article in articles:
                override = overrides.get(article.id)
                is_custom = bool(override and (override.unit_price_eco or override.unit_price_standard or override.unit_price_premium or override.designation_override))
                
                ws.cell(row=row, column=1, value=article.code).border = border
                ws.cell(row=row, column=2, value=article.category).border = border
                ws.cell(row=row, column=3, value=article.subcategory or '').border = border
                ws.cell(row=row, column=4, value=self.get_article_designation(article)).border = border
                ws.cell(row=row, column=5, value=article.unit).border = border
                ws.cell(row=row, column=6, value=float(self.get_article_price(article, 'ECO') or 0)).border = border
                ws.cell(row=row, column=7, value=float(self.get_article_price(article, 'STD') or 0)).border = border
                ws.cell(row=row, column=8, value=float(self.get_article_price(article, 'PREM') or 0)).border = border
                ws.cell(row=row, column=9, value='Oui' if is_custom else '').border = border
                
                for col in [6, 7, 8]:
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
                
                row += 1
        
        custom_articles = self.company.custom_articles.filter_by(is_active=True).all()
        
        if custom_articles:
            row += 1
            ws.cell(row=row, column=1, value="ARTICLES PERSONNALISÉS").font = Font(bold=True)
            row += 1
            
            for article in custom_articles:
                ws.cell(row=row, column=1, value=article.code).border = border
                ws.cell(row=row, column=2, value=article.category).border = border
                ws.cell(row=row, column=3, value=article.subcategory or '').border = border
                ws.cell(row=row, column=4, value=article.designation).border = border
                ws.cell(row=row, column=5, value=article.unit).border = border
                ws.cell(row=row, column=6, value=float(article.unit_price_eco or 0)).border = border
                ws.cell(row=row, column=7, value=float(article.unit_price_standard or 0)).border = border
                ws.cell(row=row, column=8, value=float(article.unit_price_premium or 0)).border = border
                ws.cell(row=row, column=9, value='Perso').border = border
                
                row += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 12
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def import_from_excel(self, file):
        try:
            wb = load_workbook(file)
            ws = wb.active
            
            stats = {'created': 0, 'updated': 0, 'errors': 0}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                
                code = str(row[0]).strip()
                category = str(row[1]).strip() if row[1] else ''
                subcategory = str(row[2]).strip() if row[2] else ''
                designation = str(row[3]).strip() if row[3] else ''
                unit = str(row[4]).strip() if row[4] else ''
                price_eco = float(row[5]) if row[5] else None
                price_std = float(row[6]) if row[6] else None
                price_prem = float(row[7]) if row[7] else None
                
                if not all([code, category, designation, unit]):
                    stats['errors'] += 1
                    continue
                
                existing = CompanyBPUArticle.query.filter_by(
                    company_id=self.company.id,
                    code=code
                ).first()
                
                if existing:
                    existing.category = category
                    existing.subcategory = subcategory
                    existing.designation = designation
                    existing.unit = unit
                    existing.unit_price_eco = price_eco
                    existing.unit_price_standard = price_std
                    existing.unit_price_premium = price_prem
                    stats['updated'] += 1
                else:
                    article = CompanyBPUArticle(
                        company_id=self.company.id,
                        code=code,
                        category=category,
                        subcategory=subcategory,
                        designation=designation,
                        unit=unit,
                        unit_price_eco=price_eco,
                        unit_price_standard=price_std,
                        unit_price_premium=price_prem
                    )
                    db.session.add(article)
                    stats['created'] += 1
            
            db.session.commit()
            
            return {'success': True, 'stats': stats}
        
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'error': str(e)}
